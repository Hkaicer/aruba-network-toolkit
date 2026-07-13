#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Aruba AP Down Monitor — hourly scheduled script (single-run)

Behavior
--------
- You schedule it to run every X minutes/hours via cron (Linux) or Task Scheduler (Windows).
- Each run:
  1) Connect via SSH (Netmiko) to one or more Aruba controllers.
  2) Run command: show ap database long status down
  3) Parse all "AP Database" sections from output (ignores headers, flags, footers, pager lines).
  4) Deduplicate rows (same AP only counted once).
  5) Update rolling history (24h and 7d JSON).
  6) If data is found:
       - Write today's CSV snapshot (ap_down_YYYY-MM-DD.csv).
       - If current number of APs down >= ALERT_THRESHOLD:
            * Check ALERT_COOLDOWN_MINUTES.
            * If cooldown expired → send instant alert email with CSV attached.
  7) At DAILY_HOUR:DAILY_MINUTE (local time):
       - Send a Daily Report (once per day) with counts and the latest snapshot.

Key safeguards
--------------
- If **no controller** yields data → no CSV and no alert.
- Parser only accepts rows with Status in {"Down", "Up"} and at least one identity (Name, MAC, Serial, or IP).
- Deduplication per controller + global deduplication.
- Alert cooldown prevents repeated spam (sends at most one alert per cooldown window).

Requirements
------------
pip install netmiko python-dotenv

.env (example)
--------------
CONTROLLERS=192.168.100.100
SSH_USERNAME=admin
SSH_PASSWORD=yourpassword
DEVICE_TYPE=aruba_os
DATA_DIR=./data
ALERT_THRESHOLD=5
ALERT_COOLDOWN_MINUTES=60
DAILY_HOUR=17
DAILY_MINUTE=0

SMTP_HOST=smtp.office365.com
SMTP_PORT=587
SMTP_USER=you@example.com
SMTP_PASSWORD=app_or_mail_password
SMTP_FROM=Aruba Monitor <you@example.com>
SMTP_TO=netops@example.com, noc@example.com
"""


import argparse
import csv
import json
import os
import re
import smtplib
import sys
import time
from datetime import datetime, timedelta, timezone, time as dt_time
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from netmiko import ConnectHandler
from dotenv import load_dotenv
import logging
from logging.handlers import RotatingFileHandler

# --- Logging setup ---
LOG_FILE = "AP_Log.log"
LOG_MAX_BYTES = 5 * 1024 * 1024  # 5 MB
LOG_BACKUP_COUNT = 5             # keep last 5 files

logger = logging.getLogger()
logger.setLevel(logging.INFO)

file_handler = RotatingFileHandler(
    LOG_FILE, maxBytes=LOG_MAX_BYTES, backupCount=LOG_BACKUP_COUNT, encoding="utf-8"
)
fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
file_handler.setFormatter(fmt)

console_handler = logging.StreamHandler(sys.stdout)
console_handler.setFormatter(fmt)

logger.handlers = []
logger.addHandler(file_handler)
logger.addHandler(console_handler)

logging.getLogger("paramiko").setLevel(logging.WARNING)
logging.getLogger("netmiko").setLevel(logging.WARNING)
# optional: sometimes 'ncclient' can be noisy too
logging.getLogger("ncclient").setLevel(logging.WARNING)


# Logging config
#LOG_FILE = "AP_Log.log"
#logging.basicConfig(
#    level=logging.INFO,
#    format="%(asctime)s [%(levelname)s] %(message)s",
#    handlers=[
#        logging.FileHandler(LOG_FILE, encoding="utf-8"),
#        logging.StreamHandler(sys.stdout)  # keep showing logs in console too
#    ]
#)
# .env support

COMMAND = "show ap database long status down"
DESIRED_COLUMNS = ["Controller", "Name", "IP Address", "Status", "Wired MAC Address", "Serial #"]

ANSI_ESCAPE_RE = re.compile(r"\x1B\[[0-?]*[ -/]*[@-~]")
CONTROL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")  # keep \x09, \x0a, \x0d

HEADER_SPLIT_RE = re.compile(r"(?:\t+|\s{2,})")
DASH_LINE_RE = re.compile(r"^\s*-{3,}\s*$")
PAGER_LINE_RE = re.compile(r"--More--", re.IGNORECASE)
APDB_HEADER_MARKER_RE = re.compile(r"\bAP\s*Database\b", re.IGNORECASE)

HEADER_SYNONYMS = {
    "name": "Name",
    "ap name": "Name",
    "ap": "Name",
    "ip address": "IP Address",
    "ip": "IP Address",
    "ipaddr": "IP Address",
    "ip-address": "IP Address",
    "status": "Status",
    "wired mac address": "Wired MAC Address",
    "wired mac": "Wired MAC Address",
    "wired-mac": "Wired MAC Address",
    "wired mac addr": "Wired MAC Address",
    "wired mac addr.": "Wired MAC Address",
    "serial": "Serial #",
    "serial #": "Serial #",
    "serial number": "Serial #",
    "serial no": "Serial #",
    "serialnum": "Serial #",
}

def parse_host_port(controller_entry: str, default_port: int) -> Tuple[str, int]:
    s = (controller_entry or "").strip()
    if not s:
        return "", default_port
    if ":" in s:
        host, port_str = s.rsplit(":", 1)
        try:
            port = int(port_str)
        except Exception:
            port = default_port
        return host.strip(), port
    return s, default_port

def now_local() -> datetime:
    return datetime.now()

def ensure_data_dir(path: Path) -> None:
    path.mkdir(parents=True, exist_ok=True)

def sanitize_text(s: str) -> str:
    s = ANSI_ESCAPE_RE.sub("", s)
    s = CONTROL_RE.sub("", s)
    s = s.replace("\r", "")
    return s

def interpret_status(status: str) -> str:
    st = (status or "").strip().lower()
    if st.startswith("down"):
        return "Down"
    if st.startswith("up"):
        return "Up"
    return status.strip() if status else ""

def parse_recipients(s: str) -> List[str]:
    if not s:
        return []
    return [p for p in re.split(r"[;,]\s*", s.strip()) if p]

def csv_today_path(data_dir: Path) -> Path:
    return data_dir / f"ap_down_{now_local():%Y-%m-%d}.csv"

def hist_24h_path(data_dir: Path) -> Path:
    return data_dir / "down_history_24h.json"

def hist_7d_path(data_dir: Path) -> Path:
    return data_dir / "down_history_7d.json"

def daily_state_path(data_dir: Path) -> Path:
    return data_dir / "daily_state.json"

def alert_state_path(data_dir: Path) -> Path:
    return data_dir / "alert_state.json"

def load_json_list(path: Path) -> List[Dict]:
    if not path.exists():
        return []
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        print(f"[WARN] Failed to parse JSON {path}: {e}", file=sys.stderr)
        return []

def save_json(path: Path, data) -> None:
    try:
        path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
    except Exception as e:
        print(f"[WARN] Failed to write JSON {path}: {e}", file=sys.stderr)

def identity_key(controller_ip: str, row: Dict[str, str]) -> str:
    for key in ["Name", "Wired MAC Address", "Serial #", "IP Address"]:
        v = (row.get(key) or "").strip()
        if v:
            return f"{controller_ip}::{v}"
    return f"{controller_ip}::unknown"

def _escape(s: str) -> str:
    return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

def _kpi_chip(label: str, value: str) -> str:
    return (
        "<div style='display:inline-block;margin:0 10px 10px 0;padding:8px 12px;"
        "border-radius:10px;background:#f5f7fa;border:1px solid #e6ebf2;"
        "font-family:Segoe UI,Arial,sans-serif;font-size:13px;'>"
        f"<span style='color:#6b7280;font-weight:600;margin-right:8px'>{_escape(label)}:</span>"
        f"<span style='color:#111827;font-weight:700'>{_escape(str(value))}</span>"
        "</div>"
    )

def _controllers_html(controllers: List[str]) -> str:
    """Always render controllers as clickable https://host:4343 links."""
    if not controllers:
        return "(none)"
    links = []
    for c in controllers:
        c = c.strip()
        if not c:
            continue
        # Always drop any existing port and force 4343
        host = c.split(":", 1)[0]
        port = "4343"
        url = f"https://{host}:{port}"
        links.append(f"<a href='{url}' style='color:#2563eb;text-decoration:none'>{host}:{port}</a>")
    return ", ".join(links)



# ---- Delta-alert identity helpers ----
from typing import Any  # at top of file you already import typing, ensure Any is imported

def ap_identity_tuple(row: Dict[str, str]) -> Tuple[str, str, str, str, str]:
    """
    Identity per controller: (Controller, Name, MAC, Serial, IP).
    Controller is required here so the same AP-name on two controllers is treated distinctly.
    """
    return (
        (row.get("Controller", "") or "").strip(),
        (row.get("Name", "") or "").strip(),
        (row.get("Wired MAC Address", "") or "").strip(),
        (row.get("Serial #", "") or "").strip(),
        (row.get("IP Address", "") or "").strip(),
    )

def load_alert_state(data_dir: Path) -> Dict[str, Any]:
    """
    alert_state.json structure:
      {
        "last_alert_sent_ts": "2025-08-29T10:12:33+00:00",
        "last_alert_keys": [
          ["Controller","Name","MAC","Serial","IP"], ...
        ]
      }
    """
    p = alert_state_path(data_dir)
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8") or "{}")
    except Exception:
        return {}

def save_alert_state(data_dir: Path, keys: List[Tuple[str, str, str, str, str]]) -> None:
    p = alert_state_path(data_dir)
    try:
        now_utc = datetime.now(timezone.utc).isoformat()
        payload = {
            "last_alert_sent_ts": now_utc,
            "last_alert_keys": list(keys),
        }
        save_json(p, payload)
    except Exception as e:
        logging.warning("Failed to write alert state: %s", e)



# -------- SSH --------

def fetch_raw_via_ssh(host: str, ssh_user: str, ssh_pass: str, device_type: str, port: int) -> Optional[str]:
    if ConnectHandler is None:
        logging.warning("Netmiko not installed; cannot SSH.")
        return None

    host = (host or "").strip()
    device = {
        "device_type": device_type or "aruba_os",
        "host": host,
        "username": ssh_user,
        "password": ssh_pass,
        "port": int(port or 22),
        "fast_cli": False,
        "global_delay_factor": 1,
    }

    try:
        with ConnectHandler(**device) as conn:
            logging.info("Connected to %s:%s via SSH", host, port)
            try:
                conn.send_command_timing("no paging", strip_command=False, strip_prompt=False)
            except Exception:
                pass

            output = ""
            page = conn.send_command_timing(COMMAND, strip_command=False, strip_prompt=False)
            output += page

            safety = 0
            while PAGER_LINE_RE.search(page or "") and safety < 200:
                page = conn.send_command_timing(" ", strip_command=False, strip_prompt=False)
                output += page
                safety += 1

            logging.info("Command executed on %s:%s", host, port)
            return output
    except Exception as e:
        logging.warning("SSH connect/exec failed to %s:%s: %s", host, port, e)
        return None


# -------- Parser --------

def _norm_label(s: str) -> str:
    s = s.lower().strip()
    s = re.sub(r"[^a-z0-9\s#-]+", "", s)
    s = re.sub(r"\s+", " ", s)
    return s

def _map_canonical(label: str) -> Optional[str]:
    return HEADER_SYNONYMS.get(_norm_label(label))

def _header_columns(line: str) -> List[Tuple[str, int, int]]:
    tokens = []
    last = 0
    for m in HEADER_SPLIT_RE.finditer(line):
        tokens.append((line[last:m.start()], last))
        last = m.end()
    tokens.append((line[last:], last))

    cols: List[Tuple[str, int, int]] = []
    for idx, (tok, start_idx) in enumerate(tokens):
        label = tok.strip()
        if not label:
            continue
        canon = _map_canonical(label)
        if not canon:
            for raw in (label, label.replace("  ", " "), label.replace("-", " ")):
                canon_try = _map_canonical(raw)
                if canon_try:
                    canon = canon_try
                    break
        if not canon:
            continue
        end_idx = tokens[idx + 1][1] if idx + 1 < len(tokens) else len(line)
        cols.append((canon, start_idx, end_idx))

    # dedupe by canonical name
    seen = set()
    out = []
    for canon, s_idx, e_idx in cols:
        if canon in seen:
            continue
        seen.add(canon)
        out.append((canon, s_idx, e_idx))
    return out

def _is_header_candidate(line: str) -> bool:
    nl = _norm_label(line)
    hits = sum(k in nl for k in ["name", "ip address", "status", "wired mac", "serial"])
    return hits >= 2

def _extract_row(line: str, cols: List[Tuple[str, int, int]]) -> Dict[str, str]:
    row: Dict[str, str] = {k: "" for k in DESIRED_COLUMNS}
    for canon, s_idx, e_idx in cols:
        cell = line[s_idx:e_idx].strip()
        if canon == "Status":
            cell = interpret_status(cell)
        row[canon] = cell
    return row

def _has_identity(row: Dict[str, str]) -> bool:
    return any((row.get(k) or "").strip() for k in ["Name", "Wired MAC Address", "Serial #", "IP Address"])

def parse_rows(raw_text: str) -> List[Dict[str, str]]:
    """
    Robust parser:
    - Accepts multiple 'AP Database' blocks
    - Ignores dashed lines, pager markers, prompts, 'Flags', 'Total APs'
    - Keeps rows ONLY if Status is 'Down' or 'Up' AND identity exists
    """
    text = sanitize_text(raw_text or "")
    if not text.strip():
        return []

    lines = text.split("\n")
    rows: List[Dict[str, str]] = []
    active_cols: Optional[List[Tuple[str, int, int]]] = None
    in_table = False

    for raw_line in lines:
        line = raw_line.rstrip()

        if not line.strip():
            continue
        if PAGER_LINE_RE.search(line):
            continue
        if DASH_LINE_RE.match(line):
            continue
        if APDB_HEADER_MARKER_RE.search(line):
            in_table = False
            active_cols = None
            continue
        if re.search(r"^\s*(flags|total aps|ap license)", line, re.IGNORECASE):
            in_table = False
            active_cols = None
            continue

        if _is_header_candidate(line):
            active_cols = _header_columns(line)
            if active_cols:
                by_name = {name: (name, s, e) for (name, s, e) in active_cols}
                ordered = []
                for want in DESIRED_COLUMNS:
                    if want in by_name:
                        ordered.append(by_name[want])
                active_cols = ordered if ordered else None
            in_table = active_cols is not None
            continue

        if in_table and active_cols:
            # Skip device prompts like "(host) *[node] #"
            if re.search(r"\)\s*#", line) or re.search(r"#\s*$", line):
                continue
            row = _extract_row(line, active_cols)
            # Accept ONLY when Status is valid and there is identity (prevents wrapped/continuation lines)
            if row.get("Status") in ("Down", "Up") and _has_identity(row):
                rows.append(row)

    return rows

# -------- CSV & History --------

def write_csv(path: Path, rows: List[Dict[str, str]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, delimiter=",")
        w.writerow(DESIRED_COLUMNS)
        for r in rows:
            w.writerow([r.get(k, "") for k in DESIRED_COLUMNS])
            
            
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def write_excel(path: Path, rows: List[Dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "AP Down"

    # Header with bold + background color
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    ws.append(DESIRED_COLUMNS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Add data rows
    for r in rows:
        ws.append([r.get(k, "") for k in DESIRED_COLUMNS])

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(path)
            




def update_histories(data_dir: Path, controller_ip: str, rows: List[Dict[str, str]]) -> None:
    now = datetime.now(timezone.utc)
    cutoff_24h = now - timedelta(hours=24)
    cutoff_7d = now - timedelta(days=7)

    p24 = hist_24h_path(data_dir)
    p7d = hist_7d_path(data_dir)

    h24 = load_json_list(p24)
    h7d = load_json_list(p7d)

    def prune(hist: List[Dict], cutoff) -> List[Dict]:
        out = []
        for e in hist:
            try:
                ts = datetime.fromisoformat(e.get("ts"))
            except Exception:
                continue
            if ts >= cutoff:
                out.append(e)
        return out

    h24 = prune(h24, cutoff_24h)
    h7d = prune(h7d, cutoff_7d)

    for r in rows:
        entry = {
            "ts": now.isoformat(),
            "controller": controller_ip,
            "key": identity_key(controller_ip, r),
            "row": {k: r.get(k, "") for k in DESIRED_COLUMNS},
        }
        h24.append(entry)
        h7d.append(entry)

    save_json(p24, h24)
    save_json(p7d, h7d)

def reconstruct_from_24h(data_dir: Path) -> List[Dict[str, str]]:
    p24 = hist_24h_path(data_dir)
    h24 = load_json_list(p24)
    latest: Dict[str, Dict] = {}
    for e in h24:
        try:
            ts = datetime.fromisoformat(e.get("ts"))
        except Exception:
            continue
        key = e.get("key")
        if not key:
            continue
        if key not in latest or ts > latest[key]["_ts"]:
            r = e.get("row", {})
            latest[key] = {"_ts": ts, **r}
    return [{k: v.get(k, "") for k in DESIRED_COLUMNS} for v in latest.values()]

def compute_unique24(data_dir: Path) -> int:
    p24 = hist_24h_path(data_dir)
    h24 = load_json_list(p24)
    return len({e.get("key") for e in h24 if e.get("key")})

# -------- Email --------

def status_badge_html(status: str) -> str:
    st = (status or "").strip().lower()
    is_down = st.startswith("down")
    color = "#e53935" if is_down else "#43a047"   # red / green
    text = "Down" if is_down else "Up"
    return (
        f'<span style="display:inline-block;padding:2px 10px;border-radius:9999px;'
        f'background:{color};color:#fff;font-weight:700;font-size:12px;">{text}</span>'
    )

def build_html(
    mode: str,                             # "ALERT" or "DAILY"
    controllers: List[str],
    rows: List[Dict[str, str]],
    unique24: int,
    title: str,
    *,
    alert_threshold: int = None,           # optional: shown on ALERT banner
    current_count: int = None,             # override rows len if needed
    cooldown_note: str = None,             # optional chip for ALERT
) -> str:
    """
    Styled HTML that clearly differentiates ALERT vs DAILY:
      - ALERT: red banner + Threshold (+ optional cooldown chip)
      - DAILY: blue banner
    """
    mode = (mode or "").upper().strip()
    is_alert = (mode == "ALERT")

    # Banner theme
    banner_bg     = "#fef2f2" if is_alert else "#eef2ff"   # pale red / pale indigo
    banner_border = "#fecaca" if is_alert else "#c7d2fe"
    banner_accent = "#b91c1c" if is_alert else "#1e3a8a"   # deep red / deep indigo
    banner_emoji  = "⚠️" if is_alert else "📊"

    controllers_html = _controllers_html(controllers)
    rows_count = len(rows) if current_count is None else int(current_count)

    # KPI chips
    kpis = []
    kpis.append(_kpi_chip("Controllers", controllers_html))
    kpis.append(_kpi_chip("Current Down", rows_count))
    kpis.append(_kpi_chip("Unique last 24h", unique24))
    if is_alert and alert_threshold is not None:
        kpis.append(_kpi_chip("Threshold", alert_threshold))
    if is_alert and cooldown_note:
        kpis.append(_kpi_chip("Cooldown", cooldown_note))

    # Table
    header_cells = "".join(
        f"<th style='text-align:left;padding:10px 12px;border-bottom:1px solid #e5e7eb;"
        f"background:#f9fafb;color:#111827;font-size:12.5px'>{_escape(h)}</th>"
        for h in DESIRED_COLUMNS
    )
    body_rows = []
    for r in rows:
        tds = []
        for k in DESIRED_COLUMNS:
            val = r.get(k, "")
            if k == "Status":
                val = status_badge_html(val)
            else:
                val = _escape(val)
            tds.append(f"<td style='padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:12.5px'>{val}</td>")
        body_rows.append("<tr>" + "".join(tds) + "</tr>")
    table_html = (
        "<table style='border-collapse:collapse;width:100%;font-family:Segoe UI,Arial,sans-serif'>"
        "<thead>" + header_cells + "</thead>"
        "<tbody>" + (''.join(body_rows) if body_rows else
                     "<tr><td colspan='5' style='padding:10px;color:#6b7280'>No rows</td></tr>") +
        "</tbody></table>"
    )

    generated = now_local().strftime("%Y-%m-%d %H:%M:%S")
    hint = 'Showing only Aruba APs reported by "show ap database long status down".'

    return f"""
    <div style="font-family:Segoe UI,Arial,sans-serif;color:#111827">
      <div style="border:1px solid {banner_border}; background:{banner_bg}; border-radius:12px; padding:14px 16px; margin-bottom:14px">
        <div style="font-size:14px; font-weight:800; color:{banner_accent}; display:flex; align-items:center;">
          <span style="font-size:18px; margin-right:8px">{banner_emoji}</span>
          <span>{_escape(title)}</span>
        </div>
        <div style="margin-top:10px">{''.join(kpis)}</div>
      </div>

      <div style="margin:10px 0 16px 0; font-size:12.5px; color:#374151">
        {_escape(hint)}
      </div>

      {table_html}

      <div style="color:#6b7280;margin-top:14px;font-size:12px">
        Generated: {generated}
      </div>
    </div>
    """
    
def build_alert_delta_html(
    controllers: List[str],
    new_rows: List[Dict[str, str]],
    still_rows: List[Dict[str, str]],
    unique24: int,
    title: str,
) -> str:
    def _escape(s: str) -> str:
        return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    def status_badge_html(status: str) -> str:
        st = (status or "").strip().lower()
        is_down = st.startswith("down")
        color = "#e53935" if is_down else "#43a047"
        text = "Down" if is_down else "Up"
        return (f'<span style="display:inline-block;padding:2px 10px;border-radius:9999px;'
                f'background:{color};color:#fff;font-weight:700;font-size:12px;">{text}</span>')

    def _table(rows: List[Dict[str, str]]) -> str:
        header_cells = "".join(
            f"<th style='text-align:left;padding:10px 12px;border-bottom:1px solid #e5e7eb;"
            f"background:#f9fafb;color:#111827;font-size:12.5px'>{_escape(h)}</th>"
            for h in DESIRED_COLUMNS
        )
        body = []
        for r in rows:
            tds = []
            for k in DESIRED_COLUMNS:
                v = r.get(k, "")
                if k == "Status":
                    v = status_badge_html(v)
                else:
                    v = _escape(v)
                tds.append(f"<td style='padding:8px 12px;border-bottom:1px solid #f3f4f6;font-size:12.5px'>{v}</td>")
            body.append("<tr>" + "".join(tds) + "</tr>")
        if not body:
            body_html = "<tr><td colspan='6' style='padding:10px;color:#6b7280'>No rows</td></tr>"
        else:
            body_html = "".join(body)
        return (
            "<table style='border-collapse:collapse;width:100%;font-family:Segoe UI,Arial,sans-serif'>"
            f"<thead>{header_cells}</thead><tbody>{body_html}</tbody></table>"
        )

    controllers_html = _controllers_html(controllers)
    generated = now_local().strftime("%Y-%m-%d %H:%M:%S")

    return f"""
    <div style="font-family:Segoe UI,Arial,sans-serif;color:#111827">
      <div style="border:1px solid #fecaca; background:#fef2f2; border-radius:12px; padding:14px 16px; margin-bottom:14px">
        <div style="font-size:14px; font-weight:800; color:#b91c1c; display:flex; align-items:center;">
          <span style="font-size:18px; margin-right:8px">⚠️</span>
          <span>{_escape(title)}</span>
        </div>
        <div style="margin-top:8px;font-size:13px;color:#374151">
          Controllers: <b>{controllers_html}</b> &nbsp;&nbsp; | &nbsp;&nbsp; Unique last 24h: <b>{unique24}</b>
        </div>
      </div>

      <h3 style="margin:16px 0 8px 0;color:#111827">🆕 New Down Since Last Alert</h3>
      {_table(new_rows)}

      <h3 style="margin:20px 0 8px 0;color:#111827">↔️ Still Down (previously reported)</h3>
      {_table(still_rows)}

      <div style="color:#6b7280;margin-top:14px;font-size:12px">
        Generated: {generated}
      </div>
    </div>
    """


def send_email(
    smtp_host: str,
    smtp_port: int,
    smtp_user: str,
    smtp_password: str,
    sender: str,
    recipients: List[str],
    subject: str,
    html_body: str,
    attachment_path: Optional[Path] = None,
) -> bool:
    if not (smtp_host and smtp_port and sender and recipients):
        print("[WARN] SMTP not fully configured; skipping email.", file=sys.stderr)
        return False
    msg = MIMEMultipart()
    msg["From"] = sender
    msg["To"] = ", ".join(recipients)
    msg["Subject"] = subject
    msg.attach(MIMEText(html_body, "html", "utf-8"))
    if attachment_path and attachment_path.exists():
        try:
            with attachment_path.open("rb") as f:
                part = MIMEBase("application", "octet-stream")
                part.set_payload(f.read())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f'attachment; filename="{attachment_path.name}"')
            msg.attach(part)
        except Exception as e:
            logging.warning("Failed attaching %s: %s", attachment_path, e)
    elif attachment_path:
        logging.warning("Attachment %s not found; sending without it.", attachment_path)

    try:
        with smtplib.SMTP(smtp_host, int(smtp_port)) as server:
            server.ehlo()
            server.starttls()
            if smtp_user and smtp_password:
                server.login(smtp_user, smtp_password)
            server.sendmail(sender, recipients, msg.as_string())
        return True
    except Exception as e:
        logging.error("Email send failed: %s", e)
        return False

# -------- Single hourly run --------

def run_once(
    controllers: List[str],
    ssh_user: str,
    ssh_pass: str,
    device_type: str,
    data_dir: Path,
    alert_threshold: int,
    daily_hour: int,
    daily_minute: int,
    smtp_conf: Dict[str, str],
    alert_cooldown_minutes: int = 0,
    default_ssh_port: int = 22,
    fallback_to_22: bool = False,
) -> int:
    ensure_data_dir(data_dir)

    # Fetch & parse per controller with per-controller dedup
    had_success = False
    all_rows: List[Dict[str, str]] = []
    all_seen_global = set()  # dedup across all controllers (by identity only)

    for ctrl in controllers:
        ctrl = ctrl.strip()
        if not ctrl:
            continue

        # Parse ip[:port] and log which port will be used
        host, port = parse_host_port(ctrl, default_port=default_ssh_port)
        if ":" in ctrl:
            logging.info("Controller %s -> using explicit port %s (from ip:port)", host, port)
        else:
            logging.info("Controller %s -> using default port %s", host, port)

        # Single SSH fetch per controller
        raw = fetch_raw_via_ssh(host, ssh_user, ssh_pass, device_type or "aruba_os", port)
        
        # --- FAILOVER/FALLBACK ---
        if raw is None and port != 22 and fallback_to_22:
            logging.info("Primary port %s failed for %s; retrying on port 22 (fallback enabled).", port, host)
            raw = fetch_raw_via_ssh(host, ssh_user, ssh_pass, device_type or "aruba_os", 22)
        # --- OPTIONAL FAILOVER END ---    
        
        if raw is None:
            continue

        rows = parse_rows(raw)

        # Per-controller dedup by identity key
        seen = set()
        dedup_rows = []
        for r in rows:
            key = identity_key(host, r)  # host is fine (IP or hostname)
            if key not in seen:
                seen.add(key)
                dedup_rows.append(r)

        logging.info("Parsed %d AP(s) from %s:%s", len(dedup_rows), host, port)
        if dedup_rows:
            had_success = True

        # Tag each row with its controller (for display + identity)
        for r in dedup_rows:
            r["Controller"] = host  # show controller as a column

        # Global dedup across all controllers (by identity WITHOUT controller so same AP isn't double-listed in one run)
        for r in dedup_rows:
            ident_no_controller = (
                (r.get("Name", "") or "").strip(),
                (r.get("Wired MAC Address", "") or "").strip(),
                (r.get("Serial #", "") or "").strip(),
                (r.get("IP Address", "") or "").strip(),
            )
            if ident_no_controller in all_seen_global:
                continue
            all_seen_global.add(ident_no_controller)
            all_rows.append(r)

    if not had_success or len(all_rows) == 0:
        logging.warning("No data fetched/parsed from any controller.")
        return 2

    # Write today's CSV (only when we had data)
    csv_path = csv_today_path(data_dir)
    write_csv(csv_path, all_rows)
    logging.info("Wrote CSV snapshot to %s", csv_path)
    
    
    excel_path = data_dir / f"ap_down_{now_local():%Y-%m-%d}.xlsx"
    write_excel(excel_path, all_rows)
    logging.info("Wrote Excel snapshot to %s", excel_path)





    # Unique in last 24h
    unique24 = compute_unique24(data_dir)

    # Build a set of identities for the CURRENT run (per-controller identity)
    current_keys = { ap_identity_tuple(r) for r in all_rows }

    # Load the last alerted identities
    astate = load_alert_state(data_dir)
    last_keys_list = astate.get("last_alert_keys", []) or []
    last_keys: set = { tuple(k) for k in last_keys_list if isinstance(k, (list, tuple)) }

    # Delta computation
    new_keys   = current_keys - last_keys
    still_keys = current_keys & last_keys

    send_alert = False
    if not last_keys:
        # First time we see any Down APs -> alert once with all as "New"
        if current_keys:
            send_alert = True
            new_keys = current_keys
            still_keys = set()
    else:
        # Subsequent runs -> only alert if there are NEW APs
        if new_keys:
            send_alert = True

    if send_alert:
        # Map key sets back to row lists for the email
        keyed_rows = { ap_identity_tuple(r): r for r in all_rows }

        new_rows = [keyed_rows[k] for k in new_keys if k in keyed_rows]
        still_rows = [keyed_rows[k] for k in still_keys if k in keyed_rows]

        logging.info("NEW AP DOWN ALERT: new=%d, still=%d, current_total=%d", len(new_rows), len(still_rows), len(current_keys))
        subj = f"[ALERT] Aruba AP Down (delta) – new:{len(new_rows)} / still:{len(still_rows)} – {now_local():%Y-%m-%d}"
        html = build_alert_delta_html(controllers, new_rows, still_rows, unique24, "Instant Alert: Aruba AP(s) Down")

        sent = send_email(
            smtp_conf.get("host", ""),
            int(smtp_conf.get("port", "587") or 587),
            smtp_conf.get("user", ""),
            smtp_conf.get("password", ""),
            smtp_conf.get("from", ""),
            smtp_conf.get("to_list", []),
            subj,
            html,
            attachment_path=excel_path,  # attach today's CSV snapshot
        )
        if sent:
            # IMPORTANT: save the current baseline ONLY when we actually sent an alert
            save_alert_state(data_dir, sorted(list(current_keys)))

    # NOTE: If there are no APs down, we DO NOT clear alert_state.json.
    # The comparison on the next run will still be against the last alert.








    # Daily report (once/day at configured time)
    try:
        state_file = daily_state_path(data_dir)
        state = {}
        if state_file.exists():
            state = json.loads(state_file.read_text(encoding="utf-8") or "{}")
        last_sent_date = state.get("last_daily_sent_date", "")

        today_str = f"{now_local():%Y-%m-%d}"
        now_time = now_local().time()
        cutoff_time = dt_time(hour=int(daily_hour), minute=int(daily_minute))

        if (now_time >= cutoff_time) and (last_sent_date != today_str):
            logging.info("Sending DAILY email...")
            daily_rows = all_rows if all_rows else reconstruct_from_24h(data_dir)
            daily_unique24 = compute_unique24(data_dir)
            subj = f"Aruba AP Daily Down – {today_str} – current:{len(daily_rows)} / unique24h:{daily_unique24}"
            title = "Daily Report: Aruba AP(s) Down"
            html = build_html(
                "DAILY",
                controllers,
                daily_rows,
                daily_unique24,
                title,
                alert_threshold=int(alert_threshold),  # optional
                current_count=len(daily_rows),
            )

            send_email(
                smtp_conf.get("host", ""),
                int(smtp_conf.get("port", "587") or 587),
                smtp_conf.get("user", ""),
                smtp_conf.get("password", ""),
                smtp_conf.get("from", ""),
                smtp_conf.get("to_list", []),
                subj,
                html,
                attachment_path=excel_path if excel_path.exists() else None,
            )
            logging.info("DAILY email sent")
            state["last_daily_sent_date"] = today_str
            save_json(state_file, state)
    except Exception as e:
        logging.warning("Daily send logic failed: %s", e)

    return 0

    # Daily report (time-gated once/day)
    try:
        state_file = daily_state_path(data_dir)
        state = {}
        if state_file.exists():
            state = json.loads(state_file.read_text(encoding="utf-8") or "{}")
        last_sent_date = state.get("last_daily_sent_date", "")

        today_str = f"{now_local():%Y-%m-%d}"
        now_time = now_local().time()
        cutoff_time = dt_time(hour=int(daily_hour), minute=int(daily_minute))

        if (now_time >= cutoff_time) and (last_sent_date != today_str):
            logging.info("Sending DAILY email...")
            daily_rows = all_rows if all_rows else reconstruct_from_24h(data_dir)
            daily_unique24 = compute_unique24(data_dir)
            subj = f"Aruba AP Daily Down – {today_str} – current:{len(daily_rows)} / unique24h:{daily_unique24}"
            title = "Daily Report: Aruba AP(s) Down"
            html = build_html(controllers, daily_rows, daily_unique24, title)
            send_email(
                smtp_conf.get("host", ""),
                int(smtp_conf.get("port", "587") or 587),
                smtp_conf.get("user", ""),
                smtp_conf.get("password", ""),
                smtp_conf.get("from", ""),
                smtp_conf.get("to_list", []),
                subj,
                html,
                attachment_path=excel_path if excel_path.exists() else None,
            )
            logging.info("DAILY email sent")
            state["last_daily_sent_date"] = today_str
            save_json(state_file, state)
    except Exception as e:
        logging.warning("Daily send logic failed: %s", e)

    return 0

# -------- CLI --------

def get_config() -> argparse.Namespace:
    load_dotenv()
    p = argparse.ArgumentParser(description="Hourly Aruba AP Down monitor: fetch via SSH, parse, update history, alert on threshold, send daily at configured time.")
    p.add_argument("--fallback-to-22",action="store_true",default=(os.getenv("FALLBACK_TO_22", "false").lower() in ("1","true","yes")),help="If set, when ssh to ip:port fails and port!=22, retry ip:22 once.")

    p.add_argument("--poll-interval-minutes",type=int,default=int(os.getenv("POLL_INTERVAL_MINUTES", "10") or 10),)
    p.add_argument("--ssh-port",type=int,default=int(os.getenv("SSH_PORT", "22")),help="Default SSH port if controller doesn't specify one (env SSH_PORT).",)
    p.add_argument("--controllers", default=os.getenv("CONTROLLERS", ""))
    p.add_argument("--ssh-username", default=os.getenv("SSH_USERNAME", ""))
    p.add_argument("--ssh-password", default=os.getenv("SSH_PASSWORD", ""))
    p.add_argument("--device-type", default=os.getenv("DEVICE_TYPE", "aruba_os"))
    p.add_argument("--data-dir", default=os.getenv("DATA_DIR", "./data"))
    p.add_argument("--alert-threshold", type=int, default=int(os.getenv("ALERT_THRESHOLD", "5") or 5))
    p.add_argument("--alert-cooldown-minutes", type=int, default=int(os.getenv("ALERT_COOLDOWN_MINUTES", "0") or 0))
    p.add_argument("--daily-hour", type=int, default=int(os.getenv("DAILY_HOUR", "17") or 17))
    p.add_argument("--daily-minute", type=int, default=int(os.getenv("DAILY_MINUTE", "0") or 0))
    p.add_argument("--smtp-host", default=os.getenv("SMTP_HOST", ""))
    p.add_argument("--smtp-port", default=os.getenv("SMTP_PORT", "587"))
    p.add_argument("--smtp-user", default=os.getenv("SMTP_USER", ""))
    p.add_argument("--smtp-password", default=os.getenv("SMTP_PASSWORD", ""))
    p.add_argument("--smtp-from", default=os.getenv("SMTP_FROM", ""))
    p.add_argument("--smtp-to", default=os.getenv("SMTP_TO", ""))
    return p.parse_args()

def main() -> int:
    args = get_config()
    controllers = [c.strip() for c in (args.controllers or "").split(",") if c.strip()]
    if not controllers:
        logging.warning("Please set CONTROLLERS (comma-separated).")
        return 2

    data_dir = Path(args.data_dir or "./data")
    ensure_data_dir(data_dir)

    smtp_conf = {
        "host": args.smtp_host,
        "port": args.smtp_port,
        "user": args.smtp_user,
        "password": args.smtp_password,
        "from": args.smtp_from,
        "to_list": parse_recipients(args.smtp_to),
    }

    # NEW: poll interval in minutes (default 10)
    poll_minutes = max(1, int(getattr(args, "poll_interval_minutes", 10)))
    logging.info("Starting continuous monitor loop. Interval=%d minutes.", poll_minutes)
    
    logging.info(
        "Default SSH port: %s (can be overridden per controller via ip:port in CONTROLLERS)",
        args.ssh_port,
    )

    try:
        while True:
            try:
                rc = run_once(
                    controllers=controllers,
                    ssh_user=args.ssh_username,
                    ssh_pass=args.ssh_password,
                    device_type=args.device_type or "aruba_os",
                    data_dir=data_dir,
                    alert_threshold=int(args.alert_threshold),
                    daily_hour=int(args.daily_hour),
                    daily_minute=int(args.daily_minute),
                    smtp_conf=smtp_conf,
                    alert_cooldown_minutes=int(args.alert_cooldown_minutes),
                    default_ssh_port=int(args.ssh_port),
                    fallback_to_22=bool(args.fallback_to_22),
                )
                logging.info("run_once finished with code %s", rc)
            except Exception as e:
                logging.error("Unhandled error in run_once: %s", e, exc_info=True)

            # Sleep between polls; 1s steps keep Ctrl+C responsive
            try:
                for _ in range(poll_minutes * 60):
                    time.sleep(1)
            except KeyboardInterrupt:
                raise

    except KeyboardInterrupt:
        logging.info("Received interrupt, exiting cleanly.")
        return 0


'''
def main() -> int:
    args = get_config()

    controllers = [c.strip() for c in (args.controllers or "").split(",") if c.strip()]
    if not controllers:
        logging.warning("Please set CONTROLLERS (comma-separated).")
        return 2

    data_dir = Path(args.data_dir or "./data")
    ensure_data_dir(data_dir)

    smtp_conf = {
        "host": args.smtp_host,
        "port": args.smtp_port,
        "user": args.smtp_user,
        "password": args.smtp_password,
        "from": args.smtp_from,
        "to_list": parse_recipients(args.smtp_to),
    }

    return run_once(
        controllers=controllers,
        ssh_user=args.ssh_username,
        ssh_pass=args.ssh_password,
        device_type=args.device_type or "aruba_os",
        data_dir=data_dir,
        alert_threshold=int(args.alert_threshold),
        daily_hour=int(args.daily_hour),
        daily_minute=int(args.daily_minute),
        smtp_conf=smtp_conf,
        alert_cooldown_minutes=int(args.alert_cooldown_minutes),
    )
'''

if __name__ == "__main__":
    sys.exit(main())
